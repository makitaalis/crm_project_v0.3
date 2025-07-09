from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, UpdateView, CreateView, View, TemplateView
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime, timedelta
from openpyxl import Workbook
from .models import Application, ApplicationHistory, ApplicationComment, Status, Category
from .forms import ApplicationForm, ApplicationOperatorForm
from apps.accounts.mixins import AdminRequiredMixin
from apps.accounts.models import User
from apps.clients.models import Client
from django.shortcuts import get_object_or_404
from datetime import timedelta

class ApplicationListView(LoginRequiredMixin, ListView):
    model = Application
    template_name = 'applications/application_list.html'
    context_object_name = 'applications'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user

        if user.is_operator():
            queryset = queryset.filter(operator=user)

        # Filters
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status__code=status)

        operator = self.request.GET.get('operator')
        if operator and user.is_admin():
            if operator == 'none':
                queryset = queryset.filter(operator__isnull=True)
            else:
                queryset = queryset.filter(operator_id=operator)

        return queryset.select_related('client', 'operator', 'status')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['statuses'] = Status.objects.all()
        if self.request.user.is_admin():
            context['operators'] = User.objects.filter(role='operator')
        return context


class ApplicationDetailView(LoginRequiredMixin, DetailView):
    model = Application
    template_name = 'applications/application_detail.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_operator():
            queryset = queryset.filter(operator=self.request.user)
        return queryset


class ApplicationAssignView(AdminRequiredMixin, UpdateView):
    model = Application
    fields = ['operator']
    template_name = 'applications/application_assign.html'

    def form_valid(self, form):
        old_operator = self.object.operator
        response = super().form_valid(form)

        # Log action
        ApplicationHistory.objects.create(
            application=self.object,
            user=self.request.user,
            action=f'Назначен оператор: {self.object.operator}'
        )

        messages.success(self.request, 'Заявка успешно назначена')
        return response

    def get_success_url(self):
        return reverse_lazy('applications:detail', kwargs={'pk': self.object.pk})


class ApplicationUpdateView(LoginRequiredMixin, UpdateView):
    model = Application
    template_name = 'applications/application_update.html'

    def get_form_class(self):
        if self.request.user.is_admin():
            return ApplicationForm
        return ApplicationOperatorForm

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.request.user.is_operator():
            queryset = queryset.filter(operator=self.request.user)
        return queryset

    def form_valid(self, form):
        old_status = self.object.status
        response = super().form_valid(form)

        # Log status change
        if 'status' in form.changed_data:
            ApplicationHistory.objects.create(
                application=self.object,
                user=self.request.user,
                action=f'Изменен статус: {old_status} → {self.object.status}',
                comment=form.cleaned_data.get('status_comment', '')
            )

        messages.success(self.request, 'Заявка успешно обновлена')
        return response

    def get_success_url(self):
        if self.request.user.is_operator():
            return reverse_lazy('applications:list')
        return reverse_lazy('applications:detail', kwargs={'pk': self.object.pk})


class ApplicationCreateView(AdminRequiredMixin, CreateView):
    model = Application
    template_name = 'applications/application_form.html'
    fields = ['client', 'operator', 'status', 'categories', 'notes']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        form.fields['client'].queryset = Client.objects.all()
        form.fields['operator'].queryset = User.objects.filter(role='operator', is_blocked=False)
        form.fields['operator'].required = False
        form.fields['status'].initial = Status.objects.filter(code='new').first()
        return form

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)

        # Логирование
        ApplicationHistory.objects.create(
            application=self.object,
            user=self.request.user,
            action='Создана заявка'
        )

        messages.success(self.request, 'Заявка успешно создана')
        return response

    def get_success_url(self):
        return reverse_lazy('applications:detail', kwargs={'pk': self.object.pk})


# Управление статусами
class StatusListView(AdminRequiredMixin, ListView):
    model = Status
    template_name = 'applications/status_list.html'
    context_object_name = 'statuses'


class StatusCreateView(AdminRequiredMixin, CreateView):
    model = Status
    fields = ['code', 'name', 'color', 'is_final', 'order']
    template_name = 'applications/status_form.html'
    success_url = reverse_lazy('applications:status_list')


class StatusUpdateView(AdminRequiredMixin, UpdateView):
    model = Status
    fields = ['name', 'color', 'is_final', 'order']
    template_name = 'applications/status_form.html'
    success_url = reverse_lazy('applications:status_list')


# Управление категориями
class CategoryListView(AdminRequiredMixin, ListView):
    model = Category
    template_name = 'applications/category_list.html'
    context_object_name = 'categories'


class CategoryCreateView(AdminRequiredMixin, CreateView):
    model = Category
    fields = ['name', 'color']
    template_name = 'applications/category_form.html'
    success_url = reverse_lazy('applications:category_list')


class CategoryUpdateView(AdminRequiredMixin, UpdateView):
    model = Category
    fields = ['name', 'color']
    template_name = 'applications/category_form.html'
    success_url = reverse_lazy('applications:category_list')


class ExportApplicationsView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Заявки"

        # Заголовки
        headers = ['ID', 'Клиент', 'Телефон', 'Email', 'Статус', 'Оператор', 'Создана', 'Обновлена', 'Примечания']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        applications = Application.objects.select_related('client', 'operator', 'status').all()
        for row, app in enumerate(applications, 2):
            ws.cell(row=row, column=1, value=app.pk)
            ws.cell(row=row, column=2, value=app.client.name or 'Без имени')
            ws.cell(row=row, column=3, value=app.client.phone or '')
            ws.cell(row=row, column=4, value=app.client.email or '')
            ws.cell(row=row, column=5, value=app.status.name)
            ws.cell(row=row, column=6, value=app.operator.get_full_name() if app.operator else 'Не назначен')
            ws.cell(row=row, column=7, value=app.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=8, value=app.updated_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=9, value=app.notes)

        # Отправляем файл
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=applications_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response


# Добавьте эти классы в конец файла apps/applications/views.py

class ApplicationMassAssignView(AdminRequiredMixin, View):
    """Массовое назначение заявок оператору"""

    def post(self, request):
        application_ids = request.POST.getlist('applications')
        operator_id = request.POST.get('operator')

        if not application_ids:
            messages.error(request, 'Выберите хотя бы одну заявку')
            return redirect('applications:list')

        if not operator_id:
            messages.error(request, 'Выберите оператора')
            return redirect('applications:list')

        try:
            operator = User.objects.get(pk=operator_id, role='operator')
            applications = Application.objects.filter(pk__in=application_ids)

            count = 0
            for app in applications:
                old_operator = app.operator
                app.operator = operator
                app.save()

                # Логирование
                ApplicationHistory.objects.create(
                    application=app,
                    user=request.user,
                    action=f'Назначен оператор: {operator.get_full_name()} (массовое назначение)'
                )
                count += 1

            messages.success(request, f'Назначено заявок: {count}')

        except User.DoesNotExist:
            messages.error(request, 'Оператор не найден')

        return redirect('applications:list')


class ApplicationQuickStatusView(LoginRequiredMixin, View):
    """Быстрая смена статуса заявки (для оператора)"""

    def post(self, request, pk):
        application = get_object_or_404(Application, pk=pk)

        # Проверка прав доступа
        if request.user.is_operator() and application.operator != request.user:
            messages.error(request, 'У вас нет прав на изменение этой заявки')
            return redirect('applications:list')

        status_code = request.POST.get('status')
        comment = request.POST.get('comment', '')

        try:
            old_status = application.status
            new_status = Status.objects.get(code=status_code)
            application.status = new_status

            # Если статус "Перезвонить", устанавливаем дату
            if status_code == 'recall':
                recall_time = request.POST.get('recall_time')
                if recall_time:
                    application.recall_date = timezone.now() + timedelta(hours=int(recall_time))

            application.save()

            # Логирование
            ApplicationHistory.objects.create(
                application=application,
                user=request.user,
                action=f'Изменен статус: {old_status.name} → {new_status.name}',
                comment=comment
            )

            # Добавляем комментарий если есть
            if comment:
                ApplicationComment.objects.create(
                    application=application,
                    user=request.user,
                    text=comment
                )

            messages.success(request, 'Статус успешно изменен')

        except Status.DoesNotExist:
            messages.error(request, 'Статус не найден')

        if request.user.is_operator():
            return redirect('applications:operator_workspace')
        return redirect('applications:detail', pk=pk)


class OperatorWorkspaceView(LoginRequiredMixin, TemplateView):
    """Рабочее место оператора"""
    template_name = 'applications/operator_workspace.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_operator():
            messages.error(request, 'Доступ только для операторов')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Мои заявки
        my_applications = Application.objects.filter(operator=user)

        # Статистика
        context['new_count'] = my_applications.filter(status__code='new').count()
        context['in_work_count'] = my_applications.filter(status__code='in_work').count()
        context['recall_count'] = my_applications.filter(status__code='recall').count()

        # Закрыто сегодня
        today = timezone.now().date()
        context['closed_today'] = my_applications.filter(
            updated_at__date=today,
            status__code__in=['success', 'fail']
        ).count()

        # Активные заявки (не закрытые)
        context['active_applications'] = my_applications.exclude(
            status__code__in=['success', 'fail']
        ).select_related('client', 'status').order_by('-created_at')

        # Заявки на перезвон, у которых подошло время
        context['recall_now'] = my_applications.filter(
            status__code='recall',
            recall_date__lte=timezone.now()
        )

        context['now'] = timezone.now()
        context['statuses'] = Status.objects.all()

        return context


# Добавьте эти классы в файл apps/applications/views.py после существующих классов

class ApplicationQuickAssignView(AdminRequiredMixin, View):
    """Быстрое назначение заявки оператору из списка"""

    def post(self, request, pk):
        application = get_object_or_404(Application, pk=pk)
        operator_id = request.POST.get('operator')

        if not operator_id:
            messages.error(request, 'Выберите оператора')
            return redirect('applications:list')

        try:
            operator = User.objects.get(pk=operator_id, role='operator', is_blocked=False)
            old_operator = application.operator

            # Назначаем оператора
            application.operator = operator
            application.save()

            # Логирование
            action = f'Назначен оператор: {operator.get_full_name()}'
            if old_operator:
                action = f'Переназначено с {old_operator.get_full_name()} на {operator.get_full_name()}'

            ApplicationHistory.objects.create(
                application=application,
                user=request.user,
                action=action
            )

            # Создаем уведомление для оператора
            ApplicationComment.objects.create(
                application=application,
                user=request.user,
                text=f'Заявка назначена вам администратором {request.user.get_full_name()}'
            )

            messages.success(request, f'Заявка #{application.pk} назначена оператору {operator.get_full_name()}')

        except User.DoesNotExist:
            messages.error(request, 'Оператор не найден')

        # Возвращаемся на страницу, с которой пришли
        next_url = request.POST.get('next', 'applications:list')
        return redirect(next_url)


class UnassignedApplicationsView(AdminRequiredMixin, ListView):
    """Отдельная страница для неназначенных заявок"""
    model = Application
    template_name = 'applications/unassigned_list.html'
    context_object_name = 'applications'
    paginate_by = 20

    def get_queryset(self):
        return Application.objects.filter(
            operator__isnull=True
        ).select_related('client', 'status').order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Получаем операторов с их текущей загрузкой
        operators = User.objects.filter(role='operator', is_blocked=False)

        for operator in operators:
            # Считаем активные заявки
            operator.active_count = Application.objects.filter(
                operator=operator,
                status__is_final=False
            ).count()
            # Определяем загруженность
            operator.workload = 'low' if operator.active_count < 5 else 'medium' if operator.active_count < 10 else 'high'

        context['operators'] = operators
        context['total_unassigned'] = self.get_queryset().count()
        return context


class OperatorStatsView(AdminRequiredMixin, TemplateView):
    """Страница со статистикой операторов для админа"""
    template_name = 'applications/operator_stats.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        operators = User.objects.filter(role='operator')

        operator_stats = []
        for operator in operators:
            stats = {
                'operator': operator,
                'total': operator.applications.count(),
                'active': operator.applications.exclude(status__is_final=True).count(),
                'today': operator.applications.filter(updated_at__date=timezone.now().date()).count(),
                'success': operator.applications.filter(status__code='success').count(),
                'fail': operator.applications.filter(status__code='fail').count(),
            }
            # Процент успешных
            if stats['success'] + stats['fail'] > 0:
                stats['success_rate'] = round(stats['success'] / (stats['success'] + stats['fail']) * 100, 1)
            else:
                stats['success_rate'] = 0

            # Статус онлайн
            stats['is_online'] = operator.last_activity >= timezone.now() - timedelta(minutes=5)

            operator_stats.append(stats)

        # Сортируем по активным заявкам
        operator_stats.sort(key=lambda x: x['active'])

        context['operator_stats'] = operator_stats
        context['total_operators'] = len(operators)
        context['online_operators'] = sum(1 for s in operator_stats if s['is_online'])

        return context


class ApplicationTransferView(AdminRequiredMixin, UpdateView):
    """Передача заявки другому оператору с комментарием"""
    model = Application
    template_name = 'applications/application_transfer.html'
    fields = []

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['operators'] = User.objects.filter(
            role='operator',
            is_blocked=False
        ).exclude(pk=self.object.operator_id if self.object.operator else None)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        new_operator_id = request.POST.get('new_operator')
        reason = request.POST.get('reason', '')

        if not new_operator_id:
            messages.error(request, 'Выберите оператора')
            return self.form_invalid(None)

        try:
            new_operator = User.objects.get(pk=new_operator_id, role='operator')
            old_operator = self.object.operator

            # Обновляем заявку
            self.object.operator = new_operator
            self.object.save()

            # Логирование
            action = f'Передана от {old_operator.get_full_name() if old_operator else "Не назначен"} к {new_operator.get_full_name()}'
            ApplicationHistory.objects.create(
                application=self.object,
                user=request.user,
                action=action,
                comment=reason
            )

            # Комментарий для нового оператора
            ApplicationComment.objects.create(
                application=self.object,
                user=request.user,
                text=f'Заявка передана вам. Причина: {reason}' if reason else 'Заявка передана вам.'
            )

            messages.success(request, f'Заявка передана оператору {new_operator.get_full_name()}')
            return redirect('applications:detail', pk=self.object.pk)

        except User.DoesNotExist:
            messages.error(request, 'Оператор не найден')
            return self.form_invalid(None)