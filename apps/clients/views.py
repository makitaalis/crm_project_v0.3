from django.db import models
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.views import View
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from .models import Client, ClientHistory
from .forms import ClientForm
from apps.accounts.mixins import AdminRequiredMixin


class ClientListView(LoginRequiredMixin, ListView):
    model = Client
    template_name = 'clients/client_list.html'
    context_object_name = 'clients'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(name__icontains=search) |
                models.Q(phone__icontains=search) |
                models.Q(email__icontains=search)
            )
        return queryset.prefetch_related('applications')


class ClientCreateView(AdminRequiredMixin, CreateView):
    model = Client
    form_class = ClientForm
    template_name = 'clients/client_form.html'
    success_url = reverse_lazy('clients:list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user

        # Проверка дубликатов
        phone = form.cleaned_data.get('phone')
        if phone and Client.objects.filter(phone=phone).exists():
            messages.warning(self.request, f'Клиент с телефоном {phone} уже существует!')

        response = super().form_valid(form)

        # Автоматически создаем заявку
        if self.request.POST.get('create_application'):
            from apps.applications.models import Application, Status, ApplicationHistory
            new_status = Status.objects.filter(code='new').first()
            if new_status:
                application = Application.objects.create(
                    client=self.object,
                    status=new_status,
                    created_by=self.request.user,
                    notes='Автоматически создана при добавлении клиента'
                )

                # Логирование создания заявки
                ApplicationHistory.objects.create(
                    application=application,
                    user=self.request.user,
                    action='Создана автоматически при добавлении клиента'
                )

                messages.info(self.request, f'Создана заявка #{application.pk} для клиента')

        messages.success(self.request, 'Клиент успешно добавлен')
        return response


class ClientUpdateView(AdminRequiredMixin, UpdateView):
    model = Client
    form_class = ClientForm
    template_name = 'clients/client_form.html'
    success_url = reverse_lazy('clients:list')

    def form_valid(self, form):
        # Логирование изменений
        original = Client.objects.get(pk=self.object.pk)
        for field in form.changed_data:
            ClientHistory.objects.create(
                client=self.object,
                user=self.request.user,
                field_name=field,
                old_value=str(getattr(original, field)),
                new_value=str(getattr(form.instance, field))
            )

        messages.success(self.request, 'Клиент успешно обновлен')
        return super().form_valid(form)


class ClientDeleteView(AdminRequiredMixin, DeleteView):
    model = Client
    template_name = 'clients/client_confirm_delete.html'
    success_url = reverse_lazy('clients:list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Клиент успешно удален')
        return super().delete(request, *args, **kwargs)


class ExportClientsView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Клиенты"

        # Заголовки колонок
        headers = ['ID', 'ФИО', 'Телефон', 'Email', 'Адрес', 'Источник',
                   'Количество заявок', 'Дата создания', 'Создал', 'Примечания']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = cell.font.copy(bold=True)

        # Получаем данные
        clients = Client.objects.select_related('created_by').prefetch_related('applications').all()

        # Заполняем данные
        for row, client in enumerate(clients, 2):
            ws.cell(row=row, column=1, value=client.pk)
            ws.cell(row=row, column=2, value=client.name or '')
            ws.cell(row=row, column=3, value=client.phone or '')
            ws.cell(row=row, column=4, value=client.email or '')
            ws.cell(row=row, column=5, value=client.address or '')
            ws.cell(row=row, column=6, value=client.source or '')
            ws.cell(row=row, column=7, value=client.applications.count())
            ws.cell(row=row, column=8, value=client.created_at.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=9, value=client.created_by.get_full_name() if client.created_by else '')
            ws.cell(row=row, column=10, value=client.notes or '')

        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # Отправляем файл
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'clients_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)

        return response